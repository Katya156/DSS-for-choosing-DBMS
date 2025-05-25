import psycopg2
import pandas as pd
import tkinter as tk
from tkinter import ttk
from data.vars import *
from dotenv import load_dotenv
import os
from collections import defaultdict
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from itertools import zip_longest

# Загружаем данные из .env файла
load_dotenv()

# Подключаемся к базе
connection = psycopg2.connect(
    host=os.getenv("DB_HOST"),
    port=os.getenv("DB_PORT"),
    dbname=os.getenv("DB_NAME"),
    user=os.getenv("DB_USER"),
    password=""
)

cursor = connection.cursor()

class OldResults(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, bg=MAINCOLOR)

        # --- СОЗДАЕМ отдельный фрейм для вывода результата ---
        self.result_frame = tk.Frame(self, bg=MAINCOLOR)
        self.result_frame.grid(row=3, column=0, columnspan=4)

        # --- Загружаем данные для комбобокса ---
        cursor.execute("""SELECT DISTINCT task_name
                          FROM results r
                          JOIN tasks t ON r.tasks_id = t.id""")
        row = cursor.fetchall()
        names = [i[0] for i in row]

        if names:
            tk.Label(self, text='Выберите задачу', font=LARGEFONT, fg=FONTCOLOR, bg=MAINCOLOR, width=30, height=3
                     ).grid(row=0, column=1, columnspan=2)
            combobox = ttk.Combobox(self, values=names, state="readonly", width=50)
            combobox.grid(row=1, column=1, columnspan=2, sticky='n')
            combobox.current(0)

            def get_result():
                # ОЧИЩАЕМ фрейм от предыдущего вывода
                for widget in self.result_frame.winfo_children():
                    widget.destroy()

                tk.Label(self.result_frame, text='Результат ранжирования', font=LARGEFONT,
                         fg=FONTCOLOR, bg=MAINCOLOR, width=30, height=3
                         ).grid(row=0, column=2, columnspan=2)

                tk.Label(self.result_frame, text='Информация о задаче', font=LARGEFONT,
                         fg=FONTCOLOR, bg=MAINCOLOR, width=30, height=3
                         ).grid(row=0, column=0, columnspan=2)

                cursor.execute(f"""SELECT dv.version_name, r.dbms_weight
                                   FROM results r
                                   JOIN tasks t ON r.tasks_id = t.id AND task_name = '{combobox.get()}'
                                   JOIN dbms_versions dv ON r.dbms_version_id = dv.id
                                   ORDER BY dbms_weight DESC""")
                row = cursor.fetchall()
                res = {}
                for i in row:
                    res[i[0]] = float(i[1])

                cursor.execute(f"""SELECT criteria_name, selected_method, task_value
                                   FROM tasks t
                                   JOIN task_info ti ON ti.tasks_id = t.id AND task_name = '{combobox.get()}'
                                   JOIN criteria c ON c.id = ti.criteria_id""")
                rows = cursor.fetchall()

                grouped_data = defaultdict(list)
                for criterion, method, value in rows:
                    grouped_data[criterion].append(value)

                tk.Label(self.result_frame, text=f"Метод ранжирования - {method}", font=SMALLFONT,
                         fg=FONTCOLOR, bg=MAINCOLOR
                         ).grid(row=1, column=0, sticky='n', columnspan=2)

                start = 2
                for criterion, values in grouped_data.items():
                    tk.Label(self.result_frame, text=f"{criterion} — {', '.join(values)}", font=SMALLFONT,
                             fg=FONTCOLOR, bg=MAINCOLOR
                             ).grid(row=start, column=0, sticky='n', columnspan=2)
                    start += 1

                start_2 = 1
                cnt = 1
                for i in res:
                    tk.Label(self.result_frame, text=f'{cnt}. {i}, вес - {res[i]}', font=SMALLFONT,
                             fg=FONTCOLOR, bg=MAINCOLOR
                             ).grid(row=start_2, column=2, sticky='n', columnspan=2)
                    start_2 += 1
                    cnt += 1

                def download_results():
                    try:
                        res_res = {'СУБД': list(res.keys()), 'Ранг': list(res.values())}
                        df_res = pd.DataFrame(res_res)
                        print(df_res)
                        data = list(zip_longest(*grouped_data.values(), fillvalue=None))
                        df_criteria = pd.DataFrame(data, columns=list(grouped_data.keys()))
                        print(df_criteria)
                        # Диалог выбора файла
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".xlsx",
                            filetypes=[("Excel файлы", "*.xlsx")],
                            title="Сохранить как..."
                        )

                        if file_path:
                            with pd.ExcelWriter(file_path) as writer:
                                df_res.to_excel(writer, sheet_name=f"DBMS Weights", index=False)
                                df_criteria.to_excel(writer, sheet_name="Criteria", index=False)

                            wb = load_workbook(file_path)
                            ws = wb["DBMS Weights"]  # Название нужного листа
                            # Вставим строку в начало (перед заголовками)
                            ws.insert_rows(1)
                            ws["A1"] = f"Метод ранжирования - {method}"
                            wb.save(file_path)

                            messagebox.showinfo("Успешно", f"Файл сохранён в:\n{file_path}")
                        else:
                            messagebox.showinfo("Отмена", "Сохранение отменено.")
                    except Exception as e:
                        messagebox.showerror("Ошибка", f"Не удалось скачать данные:\n{str(e)}")

                download_btn = tk.Button(self, text="Скачать результаты", font=MIDFONT, fg=FONTCOLOR,
                                         bg=BUTTONCOLOR, activeforeground=BUTTONCOLOR,
                                         command=download_results)
                row = max(start, start_2)
                download_btn.grid(row=row + 1, column=1, sticky='n', columnspan=2)

                tk.Button(self, text='Завершить', font=MIDFONT, bg=BUTTONCOLOR, fg=FONTCOLOR,
                          activeforeground=BUTTONCOLOR, command=controller.root.destroy
                          ).grid(row=row + 3, column=1, sticky='n', columnspan=2)

            # При нажатии кнопки
            tk.Button(self, text='Подтвердить', font=LARGEFONT, bg=BUTTONCOLOR, fg=FONTCOLOR,
                      activeforeground=BUTTONCOLOR, command=get_result).grid(row=2, column=1, columnspan=2, sticky='n')

        else:
            tk.Label(self, text='Задачи не найдены', font=LARGEFONT, fg=FONTCOLOR, bg=MAINCOLOR
                     ).grid(row=7, column=1, columnspan=2)

            tk.Button(self, text='Завершить', font=MIDFONT, bg=BUTTONCOLOR, fg=FONTCOLOR,
                      activeforeground=BUTTONCOLOR, command=controller.root.destroy
                      ).grid(row=12, column=1, sticky='n', columnspan=2)

        for i in range(4):
            self.grid_columnconfigure(i, minsize=1050/4)
        for i in range(14):
            self.grid_rowconfigure(i, minsize=50)