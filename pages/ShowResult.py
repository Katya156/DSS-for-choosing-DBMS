import psycopg2
import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import tkinter as tk
from data.vars import *
from methods.AHP import ahp
from methods.TOPSIS import topsis
from dotenv import load_dotenv
import os
from collections import defaultdict
from itertools import zip_longest

# Загружаем данные из .env файла
load_dotenv()

# Подключаемся к базе
connection = psycopg2.connect(
    host=os.getenv("DB_HOST"),
    port=os.getenv("DB_PORT"),
    dbname=os.getenv("DB_NAME"),
    user=os.getenv("DB_USER"),
    password=""
)

cursor = connection.cursor()

class ShowResult(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, bg=MAINCOLOR)
        self.controller = controller

    def set_method(self, method, criteria, criteria_out, criteria_in, task_id):

        print('Критерии внутреннего уровня:', criteria_in)
        # ОТБОР ПОДХОДЯЩИХ СУБД
        cursor.execute(f"""
            with s as (
                select version_name, cv.dbms_id, task_info_id, cv.id, ca.threshold_task
                FROM task_info c
                    JOIN tasks ca
                        ON c.tasks_id = ca.id
                        AND tasks_id = {task_id}
                    left JOIN criteria_allowed_values cav
                        ON cav.criteria_id = c.criteria_id
                        AND c.task_value = cav.allowed_value
                    left JOIN criteria_values cv
                        ON cav.id = cv.criteria_allowed_values_id
                    left JOIN dbms_versions dv
                        ON cv.dbms_id = dv.id
                where not task_value ~ '^[0-9]+(\.[0-9]+)?$'
            
                union all 
                
                select version_name, cv.dbms_id, task_info_id, cv.id, ca.threshold_task
                from task_info c
                    JOIN tasks ca
                        ON c.tasks_id = ca.id
                        AND tasks_id = {task_id}
                    left JOIN criteria_allowed_values cav
                        ON cav.criteria_id = c.criteria_id
                        and cav.allowed_value ~ '^[0-9]+(\.[0-9]+)?$'
                    left JOIN criteria_values cv
                        ON cav.id = cv.criteria_allowed_values_id
                        and cv.criteria_value ~ '^[0-9]+(\.[0-9]+)?$'
                        AND (
                            (c.comparison = '<' AND CAST(c.task_value AS NUMERIC) > CAST(cv.criteria_value AS NUMERIC))
                            OR (c.comparison = '>' AND CAST(c.task_value AS NUMERIC) < CAST(cv.criteria_value AS NUMERIC))
                        )
                    left JOIN dbms_versions dv
                        ON cv.dbms_id = dv.id
                where task_value ~ '^[0-9]+(\.[0-9]+)?$'
            ),
            s1 as (
                select count(distinct task_info_id) as cnt
                from s
            )
            SELECT version_name, dbms_id, threshold_task, (select cnt from s1) as cnt, count(distinct id)
            FROM s
            group by 1,2,3,4
            having (select cnt from s1) = count(distinct id)""")
        # Извлечение данных
        rows = cursor.fetchall()
        names = {}
        for i in rows:
            name = i[0]
            id = int(i[1])
            threshold = int(i[2])
            names[name] = id
        print(names)

        if names:

            if method == 'Метод TOPSIS':
                res = topsis(criteria, names)
            elif method == 'Метод анализa иерархий':
                res = ahp(criteria_in, criteria_out, names)

            cursor.execute(f"""select criteria_name, task_value
                                                from tasks t
                                                    join task_info ti
                                                       on ti.tasks_id = t.id
                                                       and tasks_id = '{task_id}'
                                                    join criteria c
                                                        on c.id = ti.criteria_id""")
            rows = cursor.fetchall()
            grouped_data = defaultdict(list)
            for row in rows:
                criterion, value = row
                grouped_data[criterion].append(value)

            # Вывод результата

            tk.Label(self, text='Информация о задаче', font=MIDFONT, fg=FONTCOLOR, bg=MAINCOLOR,
                     ).grid(row=1, column=0,  sticky='n', columnspan=2)

            tk.Label(self, text='Результат ранжирования', font=MIDFONT, fg=FONTCOLOR,
                     bg=MAINCOLOR
                     ).grid(row=1, column=2, sticky='n', columnspan=2)

            tk.Label(self, text=f"Метод ранжирования - {method}", font=SMALLFONT, fg=FONTCOLOR,
                     bg=MAINCOLOR
                     ).grid(row=2, column=0, sticky='n', columnspan=2)
            start = 3

            for criterion, values in grouped_data.items():
                tk.Label(self, text=f"{criterion} — {', '.join(values)}", font=SMALLFONT, fg=FONTCOLOR,
                         bg=MAINCOLOR
                         ).grid(row=start, column=0, sticky='n', columnspan=2)
                start += 1

            start_2 = 2
            cnt = 1
            for i in res:
                tk.Label(self, text=f'{cnt}. {i}, вес - {round(res[i], 3)}', font=MIDFONT, fg=FONTCOLOR,
                         bg=MAINCOLOR
                         ).grid(row=start_2, column=2, sticky='n', columnspan=2)
                if start_2 >= threshold:
                    break
                start_2 += 1
                cnt += 1

            def download_results():
                try:
                    res_res = {'СУБД': list(res.keys()), 'Ранг': list(res.values())}
                    df_res = pd.DataFrame(res_res)
                    print(df_res)
                    data = list(zip_longest(*grouped_data.values(), fillvalue=None))
                    df_criteria = pd.DataFrame(data, columns=list(grouped_data.keys()))
                    print(df_criteria)
                    # Диалог выбора файла
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel файлы", "*.xlsx")],
                        title="Сохранить как..."
                    )

                    if file_path:
                        with pd.ExcelWriter(file_path) as writer:
                            df_res.to_excel(writer, sheet_name=f"DBMS Weights", index=False)
                            df_criteria.to_excel(writer, sheet_name="Criteria", index=False)

                        wb = load_workbook(file_path)
                        ws = wb["DBMS Weights"]  # Название нужного листа
                        # Вставим строку в начало (перед заголовками)
                        ws.insert_rows(1)
                        ws["A1"] = f"Метод ранжирования - {method}"
                        wb.save(file_path)

                        messagebox.showinfo("Успешно", f"Файл сохранён в:\n{file_path}")
                    else:
                        messagebox.showinfo("Отмена", "Сохранение отменено.")
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось скачать данные:\n{str(e)}")

            download_btn = tk.Button(self, text="Скачать результаты", font=MIDFONT, fg=FONTCOLOR,
                                     bg=BUTTONCOLOR, activeforeground=BUTTONCOLOR,
                                     command=download_results)
            row = max(start, start_2)
            download_btn.grid(row=row + 1, column=1, sticky='n', columnspan=2)


            def end_program():
                cnt = 1
                for i in res:
                    cursor.execute(f"""insert into results(tasks_id, dbms_version_id, calculation_date, dbms_rank, dbms_weight)
                                                        values({task_id}, {names[i]}, current_date, {cnt}, {res[i]})""")
                    connection.commit()
                    cnt += 1
                self.controller.root.destroy()

            tk.Button(self, text='Завершить', font=MIDFONT, bg=BUTTONCOLOR, fg=FONTCOLOR,
                      activeforeground=BUTTONCOLOR, command=end_program
                      ).grid(row=row + 3, column=0, sticky='n', columnspan=4)
        else:
            tk.Label(self, text='Подходящих СУБД не найдено', font=LARGEFONT, fg=FONTCOLOR,
                     bg=MAINCOLOR
                     ).grid(row=7, column=0, sticky='n', columnspan=3)

            tk.Button(self, text='Завершить', font=MIDFONT, bg=BUTTONCOLOR, fg=FONTCOLOR,
                      activeforeground=BUTTONCOLOR, command=self.controller.root.destroy
                      ).grid(row=13, column=0, sticky='n', columnspan=3)
        for i in range(4):
            self.grid_columnconfigure(i, minsize=1050/4)
        for i in range(15):
            self.grid_rowconfigure(i, minsize=700 / 15)